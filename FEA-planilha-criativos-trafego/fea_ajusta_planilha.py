# -*- coding: utf-8 -*-
"""FEA - reorganiza 'Criativos para Tráfego.xlsx' por MÊS DE INÍCIO DA CAPTAÇÃO.

1) Move o bloco 'Elite Injector Congress 11/26' da aba Novembro26 para Setembro26
   (captação iniciou em 01/09/2026).
2) Cria na aba Setembro26 o bloco '[BR] MTI - Masterclass em Intercorrências'
   (captação inicia em 07/09/2026), com os mesmos campos e links do Drive.
3) Deixa a aba Novembro26 limpa (apenas os dias 01/11 a 30/11).
"""
import datetime as dt
from copy import copy
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

SRC, OUT = 'orig.xlsx', 'FEA-Criativos-para-Trafego-v2.xlsx'
COLS = 'ABCDEFGH'

wb = openpyxl.load_workbook(SRC)
set_ws, nov_ws = wb['Setembro26'], wb['Novembro26']


def snapshot(ws, r):
    """Guarda valor, estilo e link de uma linha inteira (A..H)."""
    out = {}
    for col in COLS:
        c = ws[f'{col}{r}']
        out[col] = {
            'v': c.value,
            's': copy(c._style),
            'hl': copy(c.hyperlink) if c.hyperlink else None,
        }
    return out


def apply_row(ws, r, data):
    for col in COLS:
        c = ws[f'{col}{r}']
        d = data.get(col)
        if d is None:
            c.value = None
            c.hyperlink = None
            continue
        c.value = d['v']
        c._style = copy(d['s'])
        c.hyperlink = None
        if d['hl'] is not None:
            hl = copy(d['hl'])
            hl.ref = c.coordinate
            c.hyperlink = hl
    ws.row_dimensions[r].height = 25.5


def blank_row(ws, r, style_src):
    for col in COLS:
        c = ws[f'{col}{r}']
        c.value = None
        c.hyperlink = None
        c._style = copy(style_src[col]['s'])


# ---------------------------------------------------------------- leitura
# Setembro: dias (data -> snapshot) e blocos de tarefa por dia
set_days, set_tasks = {}, {}
last_day = None
for r in range(2, set_ws.max_row + 1):
    a = set_ws[f'A{r}'].value
    if isinstance(a, dt.datetime):
        last_day = a.date()
        set_days[last_day] = snapshot(set_ws, r)
    elif a not in (None, ''):
        set_tasks.setdefault(last_day, []).append(snapshot(set_ws, r))

nov_days = {}
elite_rows = []
last_day = None
for r in range(2, nov_ws.max_row + 1):
    a = nov_ws[f'A{r}'].value
    if isinstance(a, dt.datetime):
        last_day = a.date()
        if a.month == 11:                      # só os dias de novembro ficam
            nov_days[last_day] = snapshot(nov_ws, r)
    elif a not in (None, ''):
        elite_rows.append(snapshot(nov_ws, r))

# estilos-modelo da aba Setembro
STYLE_TASK = {col: {'s': copy(set_ws[f'{col}16']._style)} for col in COLS}
STYLE_DAY = {col: {'s': copy(set_ws[f'{col}3']._style)} for col in COLS}
HDR_SET = snapshot(set_ws, 1)
HDR_NOV = snapshot(nov_ws, 1)
LINKS_IFF = snapshot(set_ws, 2)['H']           # célula de Links Úteis do IFF+10
LINKS_ELITE = snapshot(nov_ws, 2)['H']         # célula de Links Úteis do Elite

# ------------------------------------------------- blocos reorganizados
CAPT_ELITE = dt.date(2026, 9, 1)
CAPT_MTI = dt.date(2026, 9, 7)
FASES = ['Captação', 'Lembrete', 'Antecipação', 'Vendas', 'RMKT']

# Elite: descarta linha sem fase, padroniza estilo e deadline = início da captação
elite_by_fase = {}
for row in elite_rows:
    fase = row['D']['v']
    if fase:
        elite_by_fase[fase.strip()] = row

elite_block = []
for fase in FASES:
    src = elite_by_fase.get(fase)
    if src is None:
        continue
    novo = {}
    for col in COLS:
        novo[col] = {
            'v': src[col]['v'],
            's': copy(STYLE_TASK[col]['s']),
            'hl': src[col]['hl'],
        }
    novo['E']['v'] = dt.datetime(CAPT_ELITE.year, CAPT_ELITE.month, CAPT_ELITE.day)
    novo['H'] = {'v': None, 's': copy(STYLE_TASK['H']['s']), 'hl': None}
    elite_block.append(novo)

# MTI: mesma estrutura, links das pastas do Drive da Masterclass em Intercorrências
DRIVE = 'https://drive.google.com/drive/folders/'
MTI_LEGENDAS = DRIVE + '1c84Ifsu4QDi4LsZJS1HSLJ3y3fS8SabG?usp=drive_link'
MTI_CRIATIVOS = {
    'Captação': ('CAPTAÇÃO', DRIVE + '1vNFOtJ8RomWs9wOHhbRomyDLPUAIrk7w?usp=drive_link'),
    'Lembrete': ('LEMBRETE', DRIVE + '1a4I8l-Z6N26KVYgP4bXc3ufu-9eakKCw?usp=drive_link'),
    'Antecipação': ('ANTECIPAÇÃO', DRIVE + '1rrGOG-rDixsZmrlhbzvQAl8NWpe-Pd1Z?usp=drive_link'),
    'Vendas': ('VENDAS (pasta a criar)', None),
    'RMKT': ('REMARKETING', DRIVE + '1tLoPCC59fORZjgJNC56mf9XERiw7OBlr?usp=drive_link'),
}

mti_block = []
for fase in FASES:
    rotulo, link = MTI_CRIATIVOS[fase]
    row = {col: {'v': None, 's': copy(STYLE_TASK[col]['s']), 'hl': None} for col in COLS}
    row['A']['v'] = 'Agendar 📅'
    row['B']['v'] = '[BR] MTI - Masterclass em Intercorrências'
    row['D']['v'] = fase
    row['E']['v'] = dt.datetime(CAPT_MTI.year, CAPT_MTI.month, CAPT_MTI.day)
    row['F']['v'] = 'LEGENDAS'
    row['F']['hl'] = Hyperlink(ref='', target=MTI_LEGENDAS, display=MTI_LEGENDAS)
    row['G']['v'] = rotulo
    if link:
        row['G']['hl'] = Hyperlink(ref='', target=link, display=link)
    mti_block.append(row)

set_tasks.setdefault(CAPT_ELITE, [])
set_tasks[CAPT_ELITE] = elite_block + set_tasks[CAPT_ELITE]
set_tasks.setdefault(CAPT_MTI, [])
set_tasks[CAPT_MTI] = mti_block + set_tasks[CAPT_MTI]


# ------------------------------------------------------------- escrita
def rebuild(ws, header, days, tasks, links_cells, old_max):
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    ws.data_validations.dataValidation = []

    apply_row(ws, 1, header)
    r = 2
    day_rows, task_rows = [], []
    for day in sorted(days):
        data = dict(days[day])
        data['H'] = {'v': None, 's': copy(days[day]['H']['s']), 'hl': None}
        apply_row(ws, r, data)
        day_rows.append(r)
        r += 1
        for t in tasks.get(day, []):
            apply_row(ws, r, t)
            task_rows.append(r)
            r += 1
    last = r - 1

    # limpa sobras de linhas antigas
    for rr in range(r, max(old_max, r) + 1):
        blank_row(ws, rr, STYLE_DAY)

    # coluna Links Úteis (lista independente do calendário)
    for i, cell in enumerate(links_cells):
        c = ws.cell(row=2 + i, column=8)
        c.value = cell['v']
        c._style = copy(cell['s'])
        if cell['hl'] is not None:
            hl = copy(cell['hl'])
            hl.ref = c.coordinate
            c.hyperlink = hl

    for rr in day_rows:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
    for rr in task_rows:
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
    ws.merge_cells('B1:C1')
    return task_rows, last


def ranges(rows):
    return ' '.join(f'{c}{r}' for r in rows for c in [''])  # placeholder


set_tasks_rows, set_last = rebuild(
    set_ws, HDR_SET, set_days, set_tasks,
    [LINKS_IFF, LINKS_ELITE,
     {'v': 'Links úteis - Masterclass em Intercorrências.xlsx',
      's': copy(LINKS_ELITE['s']),
      'hl': Hyperlink(ref='', target='https://docs.google.com/spreadsheets/d/18k_EmTgFEg-XJaP8eRq8DA32lMYemt09/edit',
                      display='Links úteis - Masterclass em Intercorrências.xlsx')}],
    set_ws.max_row)

nov_tasks_rows, nov_last = rebuild(
    nov_ws, HDR_NOV, nov_days, {}, [], nov_ws.max_row)

# validações de dados da aba Setembro
STATUS = '"Criar copy ✍️,Revisar 👀,Corrigir 🚫,Agendar 📅,Agendado 👍,Publicado ✅"'
CAMPANHAS = ('"[BR] Downsell FEF - julho/25,[BR] Lançamento pago IFF+10 - set/26,'
             '[BR] MTI - Masterclass em Intercorrências,Elite Injector Congress 11/26"')
dv_status = DataValidation(type='list', formula1=STATUS, allow_blank=True)
dv_camp = DataValidation(type='list', formula1=CAMPANHAS, allow_blank=True)
set_ws.add_data_validation(dv_status)
set_ws.add_data_validation(dv_camp)
for rr in set_tasks_rows:
    dv_status.add(set_ws[f'A{rr}'])
    dv_camp.add(set_ws[f'B{rr}'])

wb.save(OUT)
print('linhas de tarefa Setembro:', set_tasks_rows, 'ultima linha:', set_last)
print('Novembro ultima linha:', nov_last)
